# Login_system.py

from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from Main import Face_Recognition_System  # Importing the main system
import openpyxl

class LoginSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Login System")
        self.root.geometry("300x200")

        self.username_var = StringVar()
        self.password_var = StringVar()

        Label(self.root, text="Username:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
        Entry(self.root, textvariable=self.username_var).grid(row=0, column=1, padx=10, pady=5)

        Label(self.root, text="Password:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        Entry(self.root, textvariable=self.password_var, show="*").grid(row=1, column=1, padx=10, pady=5)

        Button(self.root, text="Login", command=self.login).grid(row=2, columnspan=2, pady=5)
        Button(self.root, text="Register", command=self.register).grid(row=3, columnspan=2, pady=5)

        if self.is_admin_logged_in():  # Check if admin is logged in
            Button(self.root, text="Approve Registrations", command=self.approve_users).grid(row=4, columnspan=2, pady=5)


    def login(self):
        username = self.username_var.get()
        password = self.password_var.get()

        # Your authentication logic here
        wb = openpyxl.load_workbook("user_details.xlsx")
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] == username and row[1] == password:
                if row[2] == "Approved":  # Check if the status is Approved
                    self.open_system()  # Open the main system if login is successful
                    return
                else:
                    messagebox.showerror("Error", "User registration pending approval")
                    return
        
        messagebox.showerror("Error", "Invalid username or password")

    def open_system(self):
        self.root.destroy()  # Close the login window
        root = Tk()
        app = Face_Recognition_System(root)
        root.mainloop()

    def register(self):
    # Open a registration window
        register_window = Toplevel(self.root)
        register_window.title("Register New User")
        register_window.geometry("400x400")

        Label(register_window, text="Enter username:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
        self.username_entry = Entry(register_window)
        self.username_entry.grid(row=0, column=1, padx=10, pady=5)

        Label(register_window, text="Enter password:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        self.password_entry = Entry(register_window, show="*")
        self.password_entry.grid(row=1, column=1, padx=10, pady=5)

        Label(register_window, text="Phone Number:").grid(row=2, column=0, padx=10, pady=5, sticky="e")
        self.phone_entry = Entry(register_window)
        self.phone_entry.grid(row=2, column=1, padx=10, pady=5)

        Label(register_window, text="Name:").grid(row=3, column=0, padx=10, pady=5, sticky="e")
        self.name_entry = Entry(register_window)
        self.name_entry.grid(row=3, column=1, padx=10, pady=5)

        Label(register_window, text="Branch:").grid(row=4, column=0, padx=10, pady=5, sticky="e")
        self.branch_entry = Entry(register_window)
        self.branch_entry.grid(row=4, column=1, padx=10, pady=5)

        Label(register_window, text="Semester:").grid(row=5, column=0, padx=10, pady=5, sticky="e")
        self.semester_entry = Entry(register_window)
        self.semester_entry.grid(row=5, column=1, padx=10, pady=5)

        Label(register_window, text="Year:").grid(row=6, column=0, padx=10, pady=5, sticky="e")
        self.year_entry = Entry(register_window)
        self.year_entry.grid(row=6, column=1, padx=10, pady=5)

        Label(register_window, text="Enrollment Number:").grid(row=7, column=0, padx=10, pady=5, sticky="e")
        self.enrollment_entry = Entry(register_window)
        self.enrollment_entry.grid(row=7, column=1, padx=10, pady=5)

        Label(register_window, text="Course:").grid(row=8, column=0, padx=10, pady=5, sticky="e")
        self.course_entry = Entry(register_window)
        self.course_entry.grid(row=8, column=1, padx=10, pady=5)

        Label(register_window, text="Email:").grid(row=9, column=0, padx=10, pady=5, sticky="e")
        self.email_entry = Entry(register_window)
        self.email_entry.grid(row=9, column=1, padx=10, pady=5)

        Button(register_window, text="Register", command=self.save_registration).grid(row=10, columnspan=2, pady=5)

    def save_registration(self):
    # Save registration details to Excel file
        username = self.username_entry.get()
        password = self.password_entry.get()
        phone = self.phone_entry.get()
        name = self.name_entry.get()
        branch = self.branch_entry.get()
        semester = self.semester_entry.get()
        year = self.year_entry.get()
        enrollment = self.enrollment_entry.get()
        course = self.course_entry.get()
        email = self.email_entry.get()

        # Append new user details to Excel file
        wb = openpyxl.load_workbook("user_details.xlsx")
        sheet = wb.active
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=username)
        sheet.cell(row=next_row, column=2, value=password)
        sheet.cell(row=next_row, column=3, value=phone)
        sheet.cell(row=next_row, column=4, value=name)
        sheet.cell(row=next_row, column=5, value=branch)
        sheet.cell(row=next_row, column=6, value=semester)
        sheet.cell(row=next_row, column=7, value=year)
        sheet.cell(row=next_row, column=8, value=enrollment)
        sheet.cell(row=next_row, column=9, value=course)
        sheet.cell(row=next_row, column=10, value=email)

        wb.save("user_details.xlsx")
        wb.close()

        messagebox.showinfo("Success", "New user registered successfully")
        self.username_entry.delete(0, END)
        self.password_entry.delete(0, END)
        self.phone_entry.delete(0, END)
        self.name_entry.delete(0, END)
        self.branch_entry.delete(0, END)
        self.semester_entry.delete(0, END)
        self.year_entry.delete(0, END)
        self.enrollment_entry.delete(0, END)
        self.course_entry.delete(0, END)
        self.email_entry.delete(0, END)

        
    def approve_users(self):
        # Implement the logic to display and approve pending registrations
        # For example:
        register_window = Toplevel(self.root)
        register_window.title("Approve Registrations")
        register_window.geometry("300x200")

        # Display pending registrations and allow admin to approve them

    def is_admin_logged_in(self):
        # Implement the logic to check if admin is logged in
        # For example, check if the username is "admin" and password is "admin123"
        if self.username_var.get() == "Aryan" and self.password_var.get() == "123":
            return True
        else:
            return False




if __name__ == "__main__":
    root = Tk()
    app = LoginSystem(root)
    root.mainloop()
